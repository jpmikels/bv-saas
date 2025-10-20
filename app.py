from flask import Flask, request, render_template, jsonify, send_file
import pandas as pd
import pdfplumber
import os
import io
import re

app = Flask(__name__)

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def sanitize_title(title: str) -> str:
    # Excel tab name rules: max 31 chars, no : \ / ? * [ ]
    bad = set(r':\/?*[]')
    clean = ''.join(c for c in title if c not in bad)
    return (clean or "Sheet")[:31]

def copy_styles(src_cell, dst_cell):
    # copy value/formula + most common styles
    dst_cell.value = src_cell.value
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = src_cell.font
    dst_cell.fill = src_cell.fill
    dst_cell.border = src_cell.border
    dst_cell.alignment = src_cell.alignment

def copy_sheet(src_ws, dst_ws):
    # cell contents + styles
    for row in src_ws.iter_rows():
        for c in row:
            dc = dst_ws.cell(row=c.row, column=c.column)
            copy_styles(c, dc)

    # merged cells
    for rng in getattr(src_ws, 'merged_cells', []):
        dst_ws.merge_cells(str(rng))

    # column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    # row heights
    for idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[idx].height = dim.height

def add_dataframe_sheet(wb: Workbook, name: str, df):
    ws = wb.create_sheet(sanitize_title(name))
    # write headers
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=j, value=str(col))
    # write rows
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)

# ---- Limits & upload folder ----
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB request limit (Cloud Run)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ---- Smart Excel header detection ----
MONTH_TOKENS = {"jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec","total"}

def parse_excel_smart(path: str, sheet_name=0):
    # Load without headers so we can detect them
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)

    def looks_like_header(row: pd.Series) -> bool:
        vals = [str(x).strip().lower() for x in row.dropna().tolist()]
        return any(any(tok in v for tok in MONTH_TOKENS) for v in vals)

    # pick header row
    header_row_idx = None
    for i in range(min(15, len(raw))):
        if looks_like_header(raw.iloc[i]):
            header_row_idx = i
            break
    if header_row_idx is None:
        header_row_idx = raw.iloc[:15].notna().sum(axis=1).idxmax()

    # Re-read with skiprows so the row after becomes header
    df = pd.read_excel(path, sheet_name=sheet_name, header=None, skiprows=header_row_idx)
    header = df.iloc[0].astype(str).str.strip().fillna("")
    df = df.iloc[1:]  # drop header row from data

    # Assign headers, drop empty/unnamed columns
    cols = pd.Index(header)
    keep = ~cols.str.match(r"^\s*$") & ~cols.str.startswith("Unnamed")
    df = df.loc[:, keep]
    cols = cols[keep]

    # Make column names unique to avoid "Reindexing only valid..." errors
    def make_unique(idx: pd.Index) -> pd.Index:
        seen = {}
        out = []
        for c in idx:
            c = str(c)
            if c in seen:
                seen[c] += 1
                out.append(f"{c}_{seen[c]}")
            else:
                seen[c] = 0
                out.append(c)
        return pd.Index(out)

    df.columns = make_unique(cols)

    # Drop fully empty rows
    df = df.dropna(axis=0, how="all")

    # Forward-fill the first visible column using position (not label)
    if df.shape[1] > 0:
        df.iloc[:, 0] = df.iloc[:, 0].ffill()

    return df.reset_index(drop=True)

# ---- PDF tables → DataFrame ----
def parse_pdf_tables(path: str) -> pd.DataFrame:
    frames = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                df = pd.DataFrame(table[1:], columns=table[0])
                frames.append(df)
    if frames:
        df = pd.concat(frames, ignore_index=True)
        df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)
        return df
    return pd.DataFrame()

# ---- Very simple sheet classifier ----
_PNL_KW = re.compile(r"(p&l|profit|income\s*statement|revenue)", re.I)
_BS_KW  = re.compile(r"(balance\s*sheet|assets|liabilities|equity)", re.I)
_CF_KW  = re.compile(r"(cash\s*flow|operating\s*activities|investing\s*activities|financing\s*activities)", re.I)

def classify_sheet(df: pd.DataFrame, fname: str) -> str:
    name = fname.lower()
    head_vals = " ".join(map(str, df.head(5).values.flatten())).lower()
    if _PNL_KW.search(name) or _PNL_KW.search(head_vals): return "P&L"
    if _BS_KW.search(name)  or _BS_KW.search(head_vals):  return "Balance Sheet"
    if _CF_KW.search(name)  or _CF_KW.search(head_vals):  return "Cash Flow"
    return "Other"

def sanitize_title(title: str) -> str:
    bad = set(r':\/?*[]'); clean = ''.join(c for c in title if c not in bad)
    return (clean or "Sheet")[:31]

def copy_styles(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = src_cell.font
    dst_cell.fill = src_cell.fill
    dst_cell.border = src_cell.border
    dst_cell.alignment = src_cell.alignment

def copy_sheet(src_ws, dst_ws):
    for row in src_ws.iter_rows():
        for c in row:
            dc = dst_ws.cell(row=c.row, column=c.column)
            copy_styles(c, dc)
    for rng in getattr(src_ws, 'merged_cells', []):
        dst_ws.merge_cells(str(rng))
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    for idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[idx].height = dim.height

def add_dataframe_sheet(wb, name, df):
    ws = wb.create_sheet(sanitize_title(name))
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=j, value=str(col))
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)

# ---- Routes ----
@app.route('/')
def index():
    return render_template('index.html')

# Multi-file upload → consolidated Excel response
@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        files = request.files.getlist('files')
        if not files:
            return jsonify(error='No files uploaded'), 400

        master_wb = Workbook()
        master_wb.remove(master_wb.active)  # drop default sheet

        for f in files:
            if not f or not f.filename: 
                continue
            fname = f.filename
            base = os.path.splitext(os.path.basename(fname))[0]
            path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            f.save(path)

            lower = fname.lower()
            if lower.endswith('.xlsx'):
                # keep formulas & styles
                src_wb = load_workbook(path, data_only=False)
                for src_ws in src_wb.worksheets:
                    dst_ws = master_wb.create_sheet(sanitize_title(f"{base}-{src_ws.title}"))
                    copy_sheet(src_ws, dst_ws)

            elif lower.endswith('.csv'):
                df = pd.read_csv(path)
                add_dataframe_sheet(master_wb, f"{base}-CSV", df)

            elif lower.endswith('.pdf'):
                import pdfplumber
                with pdfplumber.open(path) as pdf:
                    added = False
                    for i, page in enumerate(pdf.pages, 1):
                        table = page.extract_table()
                        if table:
                            df = pd.DataFrame(table[1:], columns=table[0])
                            add_dataframe_sheet(master_wb, f"{base}-p{i}", df)
                            added = True
                    if not added:
                        add_dataframe_sheet(master_wb, f"{base}-PDF", pd.DataFrame([["No tables detected"]], columns=["Info"]))
            else:
                add_dataframe_sheet(master_wb, f"{base}-UNSUPPORTED", pd.DataFrame([["Unsupported type"]], columns=["Info"]))

        from io import BytesIO
        buf = BytesIO()
        master_wb.save(buf); buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="valuation_consolidated.xlsx"
        )
    except Exception as e:
        app.logger.exception("Upload failed")
        return jsonify(error=str(e)), 500

        # Stream back as a download AND (optionally) save a copy in /uploads
        from io import BytesIO
        buf = BytesIO()
        master_wb.save(buf)
        buf.seek(0)

        # Optional local copy (ephemeral in Cloud Run)
        # with open(os.path.join(app.config['UPLOAD_FOLDER'], "valuation_consolidated.xlsx"), "wb") as out:
        #     out.write(buf.getbuffer())

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="valuation_consolidated.xlsx"
        )
    except Exception as e:
        app.logger.exception("Upload failed")
        return jsonify(error=str(e)), 500

# ---- Cloud Run entrypoint ----
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
